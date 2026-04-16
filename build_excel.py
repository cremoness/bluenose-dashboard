from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# ── Color palette ───────────────────────────────────────────────────────────
H_BG   = "1E3A5F"   # header bg
H_FG   = "FFFFFF"   # header text
DONE_BG = "E8E8E8"  # completed row bg
DONE_FG = "888888"  # completed row text
ALT_BG  = "F4F7FB"  # alternating row bg
CHK_BG  = "D4EDDA"  # checkmark cell done
PRI_R   = "FFCCCC"  # urgente bg
PRI_Y   = "FFF3CD"  # esta semana bg
PRI_G   = "D4EDDA"  # próximo mes bg

# ── Border helper ────────────────────────────────────────────────────────────
thin = Side(style="thin", color="D0D5DD")
thick_h = Side(style="medium", color="1E3A5F")

def cell_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_border():
    return Border(left=thin, right=thin, top=thick_h, bottom=thick_h)

# ── Style helpers ────────────────────────────────────────────────────────────
def style_header(cell):
    cell.font = Font(name="Arial", bold=True, color=H_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=H_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = header_border()

def style_done(cell):
    cell.font = Font(name="Arial", color=DONE_FG, size=9, strike=True)
    cell.fill = PatternFill("solid", fgColor=DONE_BG)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = cell_border()

def style_done_plain(cell):
    cell.font = Font(name="Arial", color=DONE_FG, size=9)
    cell.fill = PatternFill("solid", fgColor=DONE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = cell_border()

def style_active(cell, alt=False):
    bg = ALT_BG if alt else "FFFFFF"
    cell.font = Font(name="Arial", size=9, color="1E293B")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = cell_border()

def style_active_center(cell, alt=False):
    bg = ALT_BG if alt else "FFFFFF"
    cell.font = Font(name="Arial", size=9, color="1E293B")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = cell_border()

def style_pri(cell, pri, done=False, alt=False):
    if done:
        style_done_plain(cell)
        return
    if pri == "🔴 Urgente":
        cell.fill = PatternFill("solid", fgColor="FFCCCC")
        cell.font = Font(name="Arial", size=9, bold=True, color="7F1D1D")
    elif pri == "🟡 Esta semana":
        cell.fill = PatternFill("solid", fgColor="FFF3CD")
        cell.font = Font(name="Arial", size=9, bold=False, color="78350F")
    else:
        cell.fill = PatternFill("solid", fgColor="D4EDDA")
        cell.font = Font(name="Arial", size=9, bold=False, color="14532D")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = cell_border()

def style_check(cell, done=False):
    if done:
        cell.value = "✓"
        cell.font = Font(name="Arial", bold=True, color="155724", size=11)
        cell.fill = PatternFill("solid", fgColor=CHK_BG)
    else:
        cell.value = ""
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = cell_border()

RESP_COLORS = {
    "BlueNose":   ("DBEAFE", "1E3A8A"),
    "Cliente":    ("FCE7F3", "831843"),
    "Compartido": ("FEF3C7", "78350F"),
}

def style_resp_type(cell, rtype, done=False):
    if done:
        style_done_plain(cell)
        return
    bg, fg = RESP_COLORS.get(rtype, ("FFFFFF", "000000"))
    cell.font = Font(name="Arial", size=9, bold=True, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = cell_border()

# ── Task sheet builder ───────────────────────────────────────────────────────
TASK_COLS = ["✓", "Proyecto", "Descripción de tarea", "Responsable",
             "Resp. Type", "Prioridad", "Tiempo estimado", "Notas"]
TASK_WIDTHS = [4, 22, 58, 18, 13, 16, 16, 30]

def add_done_conditional_formatting(ws, num_cols, max_row):
    """Cuando col A = '✓', toda la fila se vuelve gris con tachado."""
    done_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    done_font = Font(name="Arial", color="888888", strike=True, size=9)
    dxf = DifferentialStyle(fill=done_fill, font=done_font)
    rule = Rule(type="expression", dxf=dxf, formula=['=$A2="✓"'])
    last_col = get_column_letter(num_cols)
    ws.conditional_formatting.add(f"A2:{last_col}{max_row}", rule)

    # Check cell (col A) also gets green bg when ✓
    chk_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    chk_font = Font(name="Arial", bold=True, color="155724", size=11)
    dxf_chk = DifferentialStyle(fill=chk_fill, font=chk_font)
    rule_chk = Rule(type="expression", dxf=dxf_chk, formula=['=$A2="✓"'])
    ws.conditional_formatting.add(f"A2:A{max_row}", rule_chk)

def add_checkmark_dropdown(ws, max_row):
    """Dropdown en col A: seleccionar ✓ o dejar vacío."""
    dv = DataValidation(type="list", formula1='"✓,"', allow_blank=True,
                        showDropDown=False, showErrorMessage=False)
    dv.sqref = f"A2:A{max_row}"
    ws.add_data_validation(dv)

def build_task_sheet(ws, tasks):
    ws.row_dimensions[1].height = 32
    for c, (col, w) in enumerate(zip(TASK_COLS, TASK_WIDTHS), 1):
        cell = ws.cell(row=1, column=c, value=col)
        style_header(cell)
        ws.column_dimensions[get_column_letter(c)].width = w

    max_row = max(len(tasks) + 1, 50)
    for r, t in enumerate(tasks, 2):
        done = t.get("done", False)
        alt  = (r % 2 == 0)
        ws.row_dimensions[r].height = 36

        style_check(ws.cell(row=r, column=1), done)

        for c, key in enumerate(["proyecto", "desc", "resp", "resp_type", "pri", "tiempo", "notas"], 2):
            val = t.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            if c == 5:   style_resp_type(cell, val, done)
            elif c == 6: style_pri(cell, val, done, alt)
            elif c == 1: style_done(cell) if done else style_active(cell, alt)
            else:
                style_done(cell) if done else style_active(cell, alt)
                if c == 4 or c == 5 or c == 6 or c == 7:
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center", wrap_text=(c==3))

    add_done_conditional_formatting(ws, len(TASK_COLS), max_row)
    add_checkmark_dropdown(ws, max_row)
    ws.freeze_panes = "A2"

# ── Ticket sheet builder ─────────────────────────────────────────────────────
TKT_COLS   = ["✓", "Cliente", "Proyecto", "Ticket", "Responsable",
              "Resp. Type", "Prioridad", "Estado", "Notas"]
TKT_WIDTHS = [4, 10, 20, 52, 14, 13, 14, 18, 30]

def build_ticket_sheet(ws, tickets):
    ws.row_dimensions[1].height = 32
    for c, (col, w) in enumerate(zip(TKT_COLS, TKT_WIDTHS), 1):
        cell = ws.cell(row=1, column=c, value=col)
        style_header(cell)
        ws.column_dimensions[get_column_letter(c)].width = w

    max_row = max(len(tickets) + 1, 50)
    for r, t in enumerate(tickets, 2):
        done = t.get("done", False)
        alt  = (r % 2 == 0)
        ws.row_dimensions[r].height = 36

        style_check(ws.cell(row=r, column=1), done)
        for c, key in enumerate(["cliente","proyecto","ticket","resp","resp_type","pri","estado","notas"], 2):
            val = t.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            if c == 6:   style_resp_type(cell, val, done)
            elif c == 7: style_pri(cell, val, done, alt)
            else:
                style_done(cell) if done else style_active(cell, alt)
                if c in (2,3,5,6,7,8):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    add_done_conditional_formatting(ws, len(TKT_COLS), max_row)
    add_checkmark_dropdown(ws, max_row)
    ws.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════

WV_PE = [
    {"proyecto":"Débito BCP","desc":"Investigar y eliminar pagos duplicados en archivo TXT de Débito BCP — cobros duplicados detectados en primera quincena de abril, monto excede valores habituales. Crítico durante fecha de cobro","resp":"Cristian + Marcos","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Hoy","notas":""},
]

WV_HND = [
    {"proyecto":"Campaña Honduras","desc":"Crear y enviar plantilla de recontacto para WhatsApp y correos a Marcos y Comunicaciones BN — 4 plantillas WA listas, 2 correos hechos, links enviados tras la reunión","resp":"Luis Gustavo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Campaña Honduras","desc":"Modificar campaña Wash — cambiar lead ads a landing page","resp":"Luis Gustavo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Campaña Honduras","desc":"Crear formulario de registro landing page Cultivando Futuros","resp":"Luis Gustavo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Campaña Honduras","desc":"Compartir proyección financiera ROI 10 años con el equipo","resp":"Luis Gustavo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Integración Express Pago","desc":"Continuar integración Express Pago - HubSpot — mapeo propiedades","resp":"René","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Campaña Honduras","desc":"Revisar proyección financiera y elaborar planeamiento numérico metas/KPIs/pipeline Honduras","resp":"Enrique","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Planeamiento HND","desc":"Pedir a Luis Gustavo el Excel de planificación Honduras y confirmar supuestos (meta = 100 donantes/$10k, 100% de origen en pauta, CPL $4.25, tasa de conversión 3.9%) — para cerrar planificación numérica","resp":"Enrique","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Hoy","notas":""},
]

WV_CL = [
    {"proyecto":"Rendiciones F2F","desc":"Poner en contacto a Enrique con el equipo para gestionar temas administrativos del proyecto Payment Platform en Chile","resp":"Juana","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Rendiciones F2F","desc":"Enviar flujo de aprobaciones a Pablo — pendiente desde reunión anterior","resp":"Gonzalo","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Rendiciones F2F","desc":"Agregar resumen de motivos de rechazo con sumatorios y filtros en la vista 'rechazos mes actual' de la plataforma — facilita disputas y ajustes contractuales con la agencia","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Rendiciones F2F","desc":"Investigar y resolver fallo de configuración en transmisión de datos VirtualPost → HubSpot — root, teléfono y fecha de nacimiento no viajan en algunos casos","resp":"Cristian + Marcos","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Rendiciones F2F","desc":"Practicar casos en la plataforma de rendiciones mientras Cristian identifica el error de VirtualPost","resp":"Ornella + Gonzalo","resp_type":"Compartido","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Rendiciones F2F","desc":"Cambiar propiedades de test a propiedades reales en HubSpot una vez validadas las rendiciones F2F en producción","resp":"Pablo + Cristian","resp_type":"BlueNose","pri":"🟢 Próximo mes","tiempo":"","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Compartir nómina corregida de HubSpot para validación de Ornella — Nóminas de marzo procesadas y compartidas por correo con Excel de resultados. 87% positivo","resp":"Cristian","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Nóminas / Cobranza","desc":"Subir nómina del 5 de abril el jueves 02/04 tras recibir respuesta del banco","resp":"Ornella","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Nóminas / Cobranza","desc":"Evaluar si usar nómina CIMA o HubSpot una vez validada la nómina corregida de Cristian","resp":"Ornella","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Realizar backfill de pagos faltantes en HubSpot usando ID y suscripción de VirtualPost — crear pagos que no se generaron por el fallo de transmisión","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Evaluar y ajustar configuración del CAPTCHA en formulario de inicio de donación — cambiar versión para resolver conflicto sin eliminar protección contra bots","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Generar listado de carritos abandonados (contactos que rellenaron formulario 1: nombre, correo, teléfono, root) para base de recuperación y acción por WhatsApp","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Estandarizar propiedad root — Resuelto al corregir problema del doble ID de SIMMA; 16.945 contactos actualizados y 458 nuevos importados","resp":"Cristian + Marcos","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Nóminas / Cobranza","desc":"Consultar a Sofía qué data de nóminas fue importada y hasta qué fecha — determinar si hay pagos de sep-dic faltantes que subir","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":"Pablo lo pidió en reunión de soporte"},
    {"proyecto":"Nóminas / Cobranza","desc":"Validar nuevamente propiedad CLCapture / Select Captor en la copia de contacto a negocio — algunos registros no están copiando el valor","resp":"Diana","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Tickets Pendientes","desc":"Investigar ticket de Diana sobre propiedad nombre de agencia que aparece sin valor","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Tickets Pendientes","desc":"Revisar y resolver ticket HubSpot #44377462606 de WV Chile","resp":"Marcos","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Comunicar a Diana que no es posible generar listado de carritos abandonados para contactos que no pasaron por el formulario — no hay datos de contacto disponibles","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Acceso Plataforma","desc":"Crear acceso en plataforma para Francisco Vallejo (gerente de Más Recursos) para revisión previa a reunión","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    # Nuevas tareas 15 abril
    {"proyecto":"Virtual Post / Importación","desc":"Importar pagos y suscripciones faltantes de VirtualPost a HubSpot y compartir reporte el lunes — ya encontrados, proceder a importación","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Lunes","notas":""},
    {"proyecto":"Virtual Post / Importación","desc":"Contactar a Andrés (ejecutivo Virtual Post) para presionar respuesta sobre desconexión 13-15 marzo — Jorge no ha respondido","resp":"Ornella","resp_type":"Blue Nose + Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Investigar y resolver duplicados en nómina de abril (~5.000 IDs vs ~2.000 estimados), coordinando con HubSpot","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Viernes","notas":""},
    {"proyecto":"Nóminas / Cobranza","desc":"Actualizar propiedades y crear negocios para contactos de agencias (archivo compartido) — ~200-300 sin match requieren decisión","resp":"Diana","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Pasarela de Pagos","desc":"Solicitar a Víctor estado final del documento de ajustes de pasarela y agregar punto de doble envío de formulario (desde integración + HubSpot)","resp":"Paola","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Pasarela de Pagos","desc":"Agregar punto de doble validación de formularios al documento con Víctor — envío desde integración como respaldo independiente de HubSpot","resp":"Cristian + Paola","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
]

WV_EC = [
    {"proyecto":"Soporte Técnico","desc":"Fechas de pago con valor incorrecto — Martha confirmó, ticket cerrado","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Soporte Técnico","desc":"WF mueve pagos — workflow corregido, todos los pagos con fecha asignada","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Soporte Técnico","desc":"Avisar por Discord que el documento de Ecuador fue actualizado","resp":"Cristian","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Soporte Técnico","desc":"Revisar por qué la cédula no se copia del contacto al pago — validado para caso Kerly (Banco Pichincha): pagos feb/mar/abr copian cédula correctamente; mapeo de banco corregido","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Soporte Técnico","desc":"Investigar y resolver portal que se resetea y no reconoce archivos cargados — respuesta urgente","resp":"Cristian","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Soporte Técnico","desc":"Implementar columnas adicionales en archivos bancarios (fecha de pago aceptado y periodo)","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Soporte Técnico","desc":"Documentar que generación automática de pagos en HubSpot es independiente del portal","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    # Nuevas tareas 15 abril
    {"proyecto":"Cobranza / Captadores","desc":"Confirmar qué propiedad usar oficialmente: 'captador' o 'captador face to face' — equipo de ventas necesita unificar reportes de captaciones por facer","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Cobranza / Pagos","desc":"Investigar pago faltante: donante con 3 pagos en Pimentes pero solo 2 registrados en HubSpot — match con historial de transacciones","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Experiencia del Donante","desc":"Revisar documento 'experiencia del donante' y dar avance antes de próxima reunión — incluye automatización de cambio de frecuencia de pago","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Viernes","notas":""},
    {"proyecto":"Garantía de Calidad","desc":"Enviar resumen por correo (a Pablo y Enrique) de tickets de garantía: puntos 1, 2, 3, 6 cerrados ✓; punto 5 pendiente (próximas respuestas contabilidad); punto 4 pendiente (inicios mayo)","resp":"Vale","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
]

WV_ES = [
    {"proyecto":"Status semanal","desc":"Confirmar a Esther si se mantienen reuniones semanales status miércoles 16:00h","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Coordinar reunión con Esther sobre motivations y designations","resp":"Sofía","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Compartir tabla con agrupación ideal de designaciones y jerarquías para construir el modelo","resp":"Esther","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Pasar base completa de importación (designaciones + motivaciones) a Sofía para sandbox","resp":"Esther","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Enviar data de motivaciones y designaciones a Pablo para su análisis","resp":"Esther","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Generar panel en sandbox con ejemplos para validación de reportes y estructura","resp":"Sofía","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Motivations y Designations","desc":"Configurar y probar formularios con campos dependientes (motivación + designación) para creación de contactos y negocios","resp":"Sofía","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Hacer pruebas para implementar dependencias de datos sin dispersión excesiva en HubSpot","resp":"Sofía","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Armar árbol condicional para evaluar mejor estructura de motivaciones y designaciones","resp":"Sofía","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Revisar y avanzar agrupamiento de ~3.000 motivaciones para reducirlas a ~300–400 términos","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Explorar y crear árboles de dependencias entre motivaciones, designaciones y tipos para sandbox y producción","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"Pablo resolvió con workflows de código personalizado","done":True},
    {"proyecto":"Motivations y Designations","desc":"Enviar archivo actualizado de datos a Sofía para trabajar con estructura real","resp":"Vicente","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Proporcionar ejemplos de informes internacionales para replicar en HubSpot","resp":"Vicente","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Motivations y Designations","desc":"Evaluar viabilidad de base de datos externa relacional para automatizar asociaciones en HubSpot — evitar miles de propiedades manuales","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Motivations y Designations","desc":"Implementar workflows en HubSpot que automaticen llenado de campos asociados (Type 1, Type 2) tras creación de registros","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"Motivations y Designations","desc":"Compartir acta con pendientes post-Semana Santa para que todos estén al día","resp":"Jesús","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Iglesias","desc":"Recordar a Sofía que responda a Nieves sobre la revisión del import de Iglesias","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Pagos y Banco","desc":"Revisar y confirmar propiedades para generación de ficheros de pagos y sincronización con banco — validar SEPA/número de cuenta","resp":"Cristian + Ángelito","resp_type":"Compartido","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Pipeline e-commerce","desc":"Enviar histórico del pipeline de e-commerce para identificar donaciones recurrentes vs iniciales","resp":"Jesús","resp_type":"Cliente","pri":"🟢 Próximo mes","tiempo":"","notas":""},
    {"proyecto":"Motivations/Designations","desc":"Conocer y documentar las automatizaciones de código personalizado creadas por Pablo — implementación ya funcionando en producción","resp":"Marcos","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"WV ES General","desc":"Enviar al cliente cuadrito de previsión a futuro (desarrollo, servidores, licencias, etc.) — acordado en reunión de España de hoy","resp":"Pablo + Ángel","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
]

UCSP = [
    {"proyecto":"Simulador de Pensión","desc":"Coordinar reunión de onboarding/recorrido de plataforma antes de dar acceso admin a BlueNose","resp":"César","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Mi Propósito","desc":"Consultar a César conformidad equipo UCSP-San Pablo con funcionalidades actuales y futuras — luz verde levantamiento requerimientos. Verificar comentarios de Renzo sobre documentos compartidos","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Mi Propósito","desc":"Confirmar a Enrique que todo ok y nada pendiente o incompleto — Claudia no dará conformidad hasta que Ventas cierre y apruebe (exige aprobación el lunes). Evitar que devuelvan responsabilidad a BlueNose","resp":"Edgardo","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Admisión Pregrado — Agente IA","desc":"Revisar diseño del Agente IA de Admisión (fase final captación, 1ra matrícula) creado por Ángel","resp":"Pablo","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Admisión Pregrado — Agente IA","desc":"Armar esquema de cómo funcionará el Agente IA de Admisión — Pablo tiene sistema base montado para ventas, admisión y matrícula; falta darle forma para caso DGA","resp":"Pablo + Ángel","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Admisión Pregrado — Agente IA","desc":"Crear todos los campos de información en HubSpot para Agente IA — campos para recibir y organizar documentación de matrícula","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Admisión Pregrado — Agente IA","desc":"Conectar campos HubSpot al agente y diseñar primera etapa del flujo — recepción y organización documentación","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Exploración Segmentos Matrícula","desc":"Coordinar propuesta de Carmen Cari sobre exploración de segmentos (matrícula incompleta, temporánea, no matriculados)","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Exploración Segmentos Matrícula","desc":"Iniciar proyecto exploración del comportamiento de segmentos específicos de estudiantes en matrículas","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Reescribir storytelling presentación Eventos Inteligentes para reunión con Claudia","resp":"Ángel","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"HubSpot SpotLight","desc":"Ajustar workflow para solo generar negocio en HubSpot cuando tipificación Atom sea MQL","resp":"Pablo","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Enviar propuesta de cotización sistema de agentes inteligentes multicanal (ventas, admisión, servicio)","resp":"Pablo","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Configurar tracking de eventos inteligentes e integrar al lead scoring — campaña abril","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Solicitar a Tom listado tipificaciones campañas outbound para ajustar workflow","resp":"Claudia","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Reunirse con Silvia y Romina para priorizar métricas y matriz de eventos inteligentes","resp":"Claudia","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Agendar reunión con Pablo para revisar implementación y limitaciones WhatsApp asesores","resp":"Claudia","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Desarrollar tabla con metas diarias y % consecución con datos de campañas HubSpot","resp":"Juanjito","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"HubSpot SpotLight","desc":"Crear propiedad nueva en HubSpot para registrar fuentes de tráfico en negocios","resp":"Janet","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Centro de Idiomas","desc":"Armar Gantt con fechas tentativas para entrevistas adicionales — Focus Group y estudios cuantitativos","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Centro de Idiomas","desc":"Enviar Gantt del proyecto Centro de Idiomas al cliente","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"CENDES","desc":"Responder correo de María Berlanga confirmando manual de flujos sin costo adicional","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"CENDES","desc":"Revisar documento comentarios de María Berlanga y coordinar reunión revisión de flujos","resp":"Marcos","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"","done":True},
    {"proyecto":"CENDES","desc":"Dar formato al manual de Marcos y enviarlo a María Berlanga","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Mi Propósito","desc":"Enviar recordatorios recurrentes a Edgardo en canal Discord para que dé estatus actualizado — Claudia Tapia confirmó que presionará a Ventas para obtener conformidad","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Mesa de Ayuda (nuevo proyecto)","desc":"Armar propuesta para optimización de Mesa de Ayuda con IA para UCSP (César Puma / Experiencia del Estudiante) — TDR recibido, presupuesto dado, abierto a 2-3 años. Propuesta año 1 sin grandes cambios estructurales","resp":"Pablo + Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Próxima semana","notas":""},
    # Nuevas tareas 15 abril — Mi Propósito
    {"proyecto":"Mi Propósito — DevOps","desc":"Consultar con Paul tiempos de respuesta estándar para solicitudes DevOps","resp":"César","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Mi Propósito — DevOps","desc":"Informar a Paul que levante ambientes con repositorios GitHub compartidos (backend + frontend ya disponibles)","resp":"César","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Mi Propósito — DevOps","desc":"Obtener aprobación formal del equipo de ventas para matriz gráfica y diseños Figma compartidos por Blue Nose","resp":"César","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Mi Propósito — DevOps","desc":"Listar pantallas nuevas y modificaciones que ventas desea ver para que Blue Nose las prepare en Figma con secuencia lógica","resp":"César","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Mi Propósito — DevOps","desc":"Gestionar con Paul permisos de creación de máquinas virtuales y accesos admin en Azure para el equipo Blue Nose","resp":"César","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Mi Propósito — Diseño","desc":"Crear mockups en Figma para funcionalidades nuevas (simulacro masivo integrado, encuestas satisfacción por herramienta) con secuencia lógica de pantallas para ventas","resp":"Edgardo","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Mi Propósito — Diseño","desc":"Crear pantallas administrativas con flujo claro (inicio → herramientas → reportes → encuestas) para aprobación de ventas","resp":"Edgardo","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Mi Propósito — DevOps","desc":"Enviar a César listado detallado de accesos y permisos necesarios para Azure (creación máquinas virtuales)","resp":"Edgardo","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    # Nuevas tareas 15 abril — Nuevas Carreras
    {"proyecto":"Nuevas Carreras — Educación Secundaria","desc":"Armar Gantt para lanzamiento de Educación Secundaria — URGENTE, debe estar lista antes del 13 de agosto (CADE Educación)","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Esta semana","notas":""},
    {"proyecto":"Nuevas Carreras — Ciencia Política","desc":"Armar Gantt para Ciencia Política — inicio de trabajo de campo mediados de mayo, aprobación final junio","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Esta semana","notas":""},
    {"proyecto":"Nuevas Carreras","desc":"Coordinar visita presencial conjunta para levantar información de las tres carreras en una sola visita (Educación Secundaria, Ciencia Política, Derecho)","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Nuevas Carreras — Educación Secundaria","desc":"Entregar brief y plan curricular de Educación Secundaria la próxima semana (aprobación de Rectorado ya realizada)","resp":"Renzo","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Nuevas Carreras — Ciencia Política","desc":"Compartir análisis competitivo y datos de Ciencias Políticas para inicio de planificación estratégica","resp":"Renzo","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Nuevas Carreras","desc":"Coordinar disponibilidad del equipo académico para entrevistas, focus groups y reuniones de levantamiento","resp":"Cintia","resp_type":"Cliente","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Nuevas Carreras","desc":"Elaborar propuesta de valor y estrategia creativa para Educación Secundaria y Ciencia Política — una vez recibido el brief de Renzo","resp":"Enrique","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
    {"proyecto":"Consultoría Sectorial","desc":"Preparar programa y agenda para consultoría sectorial UCSP — inicio en mayo, incluye reunión de kick-off y reuniones recurrentes","resp":"Enrique","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Próxima semana","notas":""},
]

UPSJB = [
    {"proyecto":"Estrategia 26-2","desc":"Planificación estratégica de campaña 26-2 — documento Excel en notebook/drive con post-mortem de 26-1, KPIs, lecciones aprendidas y drivers estratégicos. Engloba: ventas, buyer persona padre, plan a distancia, carreras y agentes IA","resp":"Enrique + Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Estrategia 26-2","desc":"Cuantificar impacto de las ineficiencias en gestión de leads de ventas — leads desatendidos, tiempo de respuesta, porcentaje de conversión perdida. Presentar a Edgar para que tome acción correctiva","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Estrategia 26-2","desc":"Convertir documento Word de análisis estratégico 'Plan a Distancia' en checklist ejecutable con IA — campos: recomendado, ejecutado, responsable, fecha","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Estrategia 26-2","desc":"Comentar a Edgar la situación de las carreras de Administración y Contabilidad — sin propuesta de valor ni estudio de mercado, no se puede tomar decisión de cierre","resp":"Enrique","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Buyer Persona Padre","desc":"Implementar buyer persona padre en HubSpot — configurar propiedades, segmentación y nurturing multicanal (WhatsApp, correo, redes, SMS) alineado a nuevos hallazgos","resp":"Pablo / Marcos / Estuardo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Buyer Persona Padre","desc":"Crear panel de medición del buyer persona padre — fit/no-fit, canales de interacción y resultados del nurturing","resp":"Pablo / Marcos","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Agente Bot / WhatsApp","desc":"Escribir a Estuardo y/o Carlos para dar seguimiento a propuesta de agente Bot y WhatsApp — si no responden por escrito, tocarlo en reunión con Edgar presente","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Agente Bot / WhatsApp","desc":"Hacer seguimiento a Pablo sobre respuesta de Estuardo/Carlos a la propuesta de agente Bot y WhatsApp","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Planeamiento 26-2","desc":"Consultar a Pablo si el cliente compartió formalmente las metas 26-2 (total y por campus) y de dónde extraer CPL y tasa de conversión para planificación numérica","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Hoy","notas":""},
    {"proyecto":"Planeamiento 26-2","desc":"Reunión post-status (briefing ~5 min) para revisar KPIs y paneles HubSpot a usar en cuantificación de impacto de leads 26-1","resp":"Ángelito + Enrique","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Hoy","notas":""},
    {"proyecto":"Buyer Persona Padre","desc":"Preguntar a Estuardo en la reunión quincenal de hoy sobre implementación del buyer persona (quién lo hace: cliente o BN) y acordar pasos concretos — si dice que lo hace después, hacer seguimiento inmediato en 2 horas","resp":"Enrique","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Hoy","notas":""},
]

AGÉNTICA = [
    {"proyecto":"Landing Page","desc":"Subir maqueta HTML de landing de Agéntica a repositorio GitHub para revisión de Pablo y Enrique","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Landing Page","desc":"Crear maqueta de landing page de AGÉNTICA — nueva unidad de BlueNose, inspirada en Conversia/Impulse, por verticales/sectores, con branding BlueNose","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"","notas":""},
    {"proyecto":"Campaña BlueNose","desc":"Preparar reporte macro y micro de la campaña BlueNose — resultados por vertical, segmento, nivel jerárquico y canal. Enrique quiere ver si funciona o no y qué ajustar","resp":"Ángelito","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Campaña BlueNose","desc":"Evaluar con Pati y Pablo contratación de HubSpot Marketing Pro — necesario para habilitar nurturing automatizado, secuencias y mayor alcance de campaña","resp":"Enrique","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
    {"proyecto":"Contenido LinkedIn","desc":"Continuar publicando posts individuales por agente en LinkedIn — misma estructura del PDF '8 agentes, cero excusas'. Coordinar con Marcos el calendario de envíos","resp":"Ángelito + Mirela","resp_type":"BlueNose","pri":"🟢 Próximo mes","tiempo":"","notas":""},
    {"proyecto":"Community Manager Agent","desc":"Finalizar y preparar para lanzamiento el agente Community Manager automatizado — casi listo, primer cliente confirmado en España (gimnasio de pilates). Prioridad alta por potencial de venta","resp":"Pablo","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Esta semana","notas":""},
    {"proyecto":"Campaña Outreach","desc":"Elaborar y enviar documento con mensajes personalizados para los 5 segmentos de la red de contactos — para activar campaña de outreach","resp":"Ángelito","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Hoy","notas":""},
    {"proyecto":"LinkedIn Carruseles","desc":"Publicar carruseles de LinkedIn por agente ya listos en Drive — coordinar calendario de publicación con Marcos","resp":"Ángelito + Mirela","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
]

CAJA_ICA = [
    {"proyecto":"Marketing Cloud (Salesforce)","desc":"Inicio de proyecto de implementación Marketing Cloud — aprobado formalmente. Trabajando con equipo desarrollador en Brasil, partners de Salesforce. Próximos pasos: Service Cloud y Cels Cloud en meses siguientes","resp":"Pablo","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"","notas":""},
]

BUENAVENTURA = [
    {"proyecto":"Onboarding IA","desc":"Proyecto de onboarding y adopción de IA aprobado informalmente — taller dos jornadas Full Day con gerente general, CEO y vicepresidentes de Buenaventura Minera. Pendiente aprobación formal","resp":"Pablo + Enrique","resp_type":"BlueNose","pri":"🟡 Esta semana","tiempo":"Completado","notas":"Formalmente confirmado por VP de Finanzas","done":True},
    {"proyecto":"Taller IA","desc":"Preparar contenido y logística del taller de IA Full Day para Buenaventura — dos jornadas con CEO, gerente general y vicepresidentes. Proyecto formalmente confirmado por VP de Finanzas","resp":"Pablo + Enrique","resp_type":"BlueNose","pri":"🔴 Urgente","tiempo":"Pendiente fecha del cliente","notas":""},
]

TICKETS = [
    {"cliente":"WV CL","proyecto":"Rendiciones F2F","ticket":"Importación de negocios — en manos de Danni","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"En progreso","notas":""},
    {"cliente":"WV CL","proyecto":"Rendiciones F2F","ticket":"Negocios F2F creados erroneamente — a la espera de validación de Danni","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"En espera cliente","notas":""},
    {"cliente":"WV CL","proyecto":"Rendiciones F2F","ticket":"Nóminas Transdata — respondido, sin respuesta del cliente","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"En espera cliente","notas":""},
    {"cliente":"WV CL","proyecto":"Rendiciones F2F","ticket":"Creación de negocios donantes únicos — solucionado, pendiente validación","resp":"Cristian","resp_type":"BlueNose","pri":"🟢 Próximo mes","estado":"En espera cliente","notas":""},
    {"cliente":"WV CL","proyecto":"Nóminas","ticket":"Propiedad nombre agencia sin valor — investigando","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"En progreso","notas":""},
    {"cliente":"WV CL","proyecto":"General","ticket":"Ticket #44377462606 — asignado a Marcos para revisión","resp":"Marcos","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"En progreso","notas":""},
    {"cliente":"WV EC","proyecto":"Soporte","ticket":"Cédula no se copia del contacto al pago — pendiente confirmar","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"En progreso","notas":""},
    {"cliente":"WV EC","proyecto":"Soporte","ticket":"Portal se resetea y no reconoce archivos cargados","resp":"Cristian","resp_type":"BlueNose","pri":"🔴 Urgente","estado":"En progreso","notas":""},
    {"cliente":"WV EC","proyecto":"Soporte","ticket":"Fechas de pago con valor incorrecto — cerrado","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"Cerrado","notas":"","done":True},
    {"cliente":"WV EC","proyecto":"Soporte","ticket":"WF mueve pagos — workflow corregido","resp":"Cristian","resp_type":"BlueNose","pri":"🟡 Esta semana","estado":"Cerrado","notas":"","done":True},
]

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

# Remove default sheet
del wb[wb.sheetnames[0]]

sheets = [
    ("WV HND",       WV_HND),
    ("WV CL",        WV_CL),
    ("WV EC",        WV_EC),
    ("WV PE",        WV_PE),
    ("WV ES",        WV_ES),
    ("UCSP",         UCSP),
    ("UPSJB",        UPSJB),
    ("AGÉNTICA",     AGÉNTICA),
    ("Caja Ica",     CAJA_ICA),
    ("Buenaventura", BUENAVENTURA),
]

for name, data in sheets:
    ws = wb.create_sheet(name)
    build_task_sheet(ws, data)

ws_tickets = wb.create_sheet("Tickets")
build_ticket_sheet(ws_tickets, TICKETS)

# ── Resumen sheet ────────────────────────────────────────────────────────────
ws_res = wb.create_sheet("Resumen", 0)
ws_res.sheet_properties.tabColor = "1E3A5F"

res_headers = ["Cliente", "Total tareas", "🔴 Urgente", "🟡 Esta semana",
               "🟢 Próximo mes", "✅ Completadas", "% Completado"]
res_widths  = [20, 14, 12, 16, 14, 14, 14]

ws_res.row_dimensions[1].height = 14
ws_res.row_dimensions[2].height = 10

# Title
ws_res.merge_cells("A1:G1")
title_cell = ws_res["A1"]
title_cell.value = "📊 BlueNose — Resumen de Tareas · Semana 15 abril 2026"
title_cell.font = Font(name="Arial", bold=True, size=13, color=H_FG)
title_cell.fill = PatternFill("solid", fgColor=H_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_res.row_dimensions[3].height = 28
for c, (h, w) in enumerate(zip(res_headers, res_widths), 1):
    cell = ws_res.cell(row=3, column=c, value=h)
    style_header(cell)
    ws_res.column_dimensions[get_column_letter(c)].width = w

all_data = {n: d for n, d in sheets}
for r, (name, data) in enumerate(sheets, 4):
    total  = len(data)
    done   = sum(1 for t in data if t.get("done"))
    urg    = sum(1 for t in data if not t.get("done") and t.get("pri") == "🔴 Urgente")
    semana = sum(1 for t in data if not t.get("done") and t.get("pri") == "🟡 Esta semana")
    mes    = sum(1 for t in data if not t.get("done") and t.get("pri") == "🟢 Próximo mes")
    pct_formula = f"=F{r}/B{r}" if total > 0 else "0"

    ws_res.row_dimensions[r].height = 24
    alt = (r % 2 == 0)
    row_vals = [name, total, urg, semana, mes, done, pct_formula]
    for c, val in enumerate(row_vals, 1):
        cell = ws_res.cell(row=r, column=c, value=val)
        bg = ALT_BG if alt else "FFFFFF"
        if c == 1:
            cell.font = Font(name="Arial", bold=True, size=10, color="1E3A5F")
            cell.fill = PatternFill("solid", fgColor=bg)
        elif c == 3 and urg > 0:
            cell.font = Font(name="Arial", bold=True, size=10, color="7F1D1D")
            cell.fill = PatternFill("solid", fgColor="FFCCCC")
        elif c == 7:
            cell.font = Font(name="Arial", size=10, color="155724")
            cell.fill = PatternFill("solid", fgColor="D4EDDA")
            cell.number_format = "0%"
        else:
            cell.font = Font(name="Arial", size=10, color="1E293B")
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        cell.border = cell_border()

ws_res.freeze_panes = "A4"

# ── Tab colors ───────────────────────────────────────────────────────────────
TAB_COLORS = {
    "WV HND":"10B981","WV CL":"0891B2","WV EC":"F59E0B","WV PE":"DC2626","WV ES":"EF4444",
    "UCSP":"3B82F6","UPSJB":"F97316","AGÉNTICA":"8B5CF6",
    "Caja Ica":"FBBF24","Buenaventura":"6366F1","Tickets":"64748B","Resumen":"1E3A5F"
}
for ws in wb.worksheets:
    color = TAB_COLORS.get(ws.title)
    if color:
        ws.sheet_properties.tabColor = color

out = r"C:\Users\angel\OneDrive\Escritorio\bluenose-dashboard\bluenose_tasks_15abril.xlsx"
wb.save(out)
print(f"Saved: {out}")
